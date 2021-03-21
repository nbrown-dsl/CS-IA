from tkinter import * 
import tkinter as tk
from tkinter import ttk 
from openpyxl.workbook import Workbook 
from openpyxl import load_workbook

root = Tk()
win = tk.Tk()
win.title("User page")
win.geometry("1000x500")



v = tk.StringVar()
def setText(word):
    v.set(word)

a = ttk.Button(win, text="Math", command=setText("Math"))
a.pack()
b = ttk.Button(win, text="English", command=setText("English"))
b.pack()
c = ttk.Entry(win, textvariable=v)
c.pack()

data = Workbook()

data = load_workbook("database.xlsx")

sheet = data.active

column_a = sheet['A']
column_b = sheet['B']

print(column_a)

def get_a(): 
    list = ''
    for cell in column_a:
        list = f'{list + str (cell.value)}\n'
        
    label_1.config(text=list)

def get_b(): 
    list = ''
    for cell in column_b:
        list = f'{list + str (cell.value)}\n'
        
    label_2.config(text=list)

Gather_A = Button(root, text = "Gather names", command=get_a)
Gather_A.pack(pady= 20)

label_1 = Label(root, text="")
label_1.pack(pady=20)

Gather_B = Button(root, text = "Gather subjects", command=get_b)
Gather_B.pack(pady= 20)

label_2 = Label(root, text="")
label_2.pack(pady=20)

win.mainloop()
root.mainloop()

#need to have pip install openpyxl installed 